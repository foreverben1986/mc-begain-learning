import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook


def generateXls(fileName, sheetName, value):
    wb = Workbook()
    ws = wb.create_sheet(sheetName)
    columnIndex = 1
    for col in value:
        rowIndex = 1
        for item in col:
            ws.cell(rowIndex, columnIndex, item)
            rowIndex = rowIndex + 1
        columnIndex = columnIndex + 1
    wb.save(fileName)