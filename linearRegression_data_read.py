import matplotlib.pyplot as plt
from openpyxl import load_workbook

def readXls(fileName, sheetName):
    wb = load_workbook(fileName)
    sheet = wb[sheetName]
    col = 1
    row = 1
    result=[]
    while sheet.cell(row,col).value != None:
        colValues = []
        while sheet.cell(row, col).value != None:
            colValues.append(sheet.cell(row, col).value)
            col = col +1
        result.append(tuple(colValues))
        col = 1
        row = row + 1
    return result